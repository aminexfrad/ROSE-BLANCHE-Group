"""
© 2025 Mohamed Amine FRAD. All rights reserved.
Unauthorized use, reproduction, or modification of this code is strictly prohibited.
Intellectual Property – Protected by international copyright law.
"""

from django.contrib import admin
from django.utils.html import format_html
from django.urls import reverse
from django.utils.safestring import mark_safe
from .models import InternKpiEvaluation
from django.utils import timezone

@admin.register(InternKpiEvaluation)
class InternKpiEvaluationAdmin(admin.ModelAdmin):
    """Interface d'administration pour les évaluations KPI des stagiaires"""
    
    list_display = [
        'intern_name', 'evaluator_name', 'stage_display', 'evaluation_date', 
        'total_score_display', 'interpretation_display', 'created_at'
    ]
    
    list_filter = [
        'interpretation', 'evaluation_date', 'created_at', 
        'intern__entreprise', 'stage__company_entreprise'
    ]
    
    search_fields = [
        'intern__nom', 'intern__prenom', 'intern__email',
        'evaluator__nom', 'evaluator__prenom', 'evaluator__email',
        'stage__title', 'comments'
    ]
    
    readonly_fields = [
        'total_score', 'interpretation', 'created_at', 'updated_at',
        'score_details_display', 'weights_summary_display'
    ]
    
    fieldsets = (
        ('Informations générales', {
            'fields': (
                'intern', 'evaluator', 'stage', 'evaluation_date'
            )
        }),
        ('Évaluation KPI', {
            'fields': (
                'delivery_satisfaction_rate', 'deadline_respect_rate', 
                'learning_capacity', 'initiative_taking', 
                'professional_behavior', 'adaptability'
            ),
            'description': 'Notez chaque KPI sur 5 selon les critères définis'
        }),
        ('Résultats calculés', {
            'fields': (
                'total_score', 'interpretation', 'score_details_display', 
                'weights_summary_display'
            ),
            'classes': ('collapse',)
        }),
        ('Commentaires et métadonnées', {
            'fields': ('comments', 'created_at', 'updated_at'),
            'classes': ('collapse',)
        })
    )
    
    def intern_name(self, obj):
        """Afficher le nom complet du stagiaire avec lien"""
        if obj.intern:
            url = reverse('admin:auth_service_user_change', args=[obj.intern.id])
            return format_html('<a href="{}">{}</a>', url, obj.intern.get_full_name())
        return '-'
    intern_name.short_description = 'Stagiaire'
    intern_name.admin_order_field = 'intern__nom'
    
    def evaluator_name(self, obj):
        """Afficher le nom complet de l'évaluateur avec lien"""
        if obj.evaluator:
            url = reverse('admin:auth_service_user_change', args=[obj.evaluator.id])
            return format_html('<a href="{}">{}</a>', url, obj.evaluator.get_full_name())
        return '-'
    evaluator_name.short_description = 'Évaluateur'
    evaluator_name.admin_order_field = 'evaluator__nom'
    
    def stage_display(self, obj):
        """Afficher le titre du stage"""
        if obj.stage:
            return obj.stage.title
        return '-'
    stage_display.short_description = 'Stage'
    stage_display.admin_order_field = 'stage__title'
    
    def total_score_display(self, obj):
        """Afficher le score total avec couleur selon l'interprétation"""
        if obj.total_score is None:
            return '-'
        
        # Définir la couleur selon l'interprétation
        if obj.interpretation == 'elevé':
            color = '#28a745'  # Vert
        elif obj.interpretation == 'bon':
            color = '#17a2b8'  # Bleu
        elif obj.interpretation == 'moyen':
            color = '#ffc107'  # Jaune
        else:  # à renforcer
            color = '#dc3545'  # Rouge
        
        return format_html(
            '<span style="color: {}; font-weight: bold;">{}/5</span>',
            color, obj.total_score
        )
    total_score_display.short_description = 'Score Total'
    total_score_display.admin_order_field = 'total_score'
    
    def interpretation_display(self, obj):
        """Afficher l'interprétation avec couleur"""
        if not obj.interpretation:
            return '-'
        
        # Définir la couleur selon l'interprétation
        color_map = {
            'elevé': '#28a745',      # Vert
            'bon': '#17a2b8',        # Bleu
            'moyen': '#ffc107',      # Jaune
            'à renforcer': '#dc3545' # Rouge
        }
        
        color = color_map.get(obj.interpretation, '#6c757d')
        return format_html(
            '<span style="color: {}; font-weight: bold;">{}</span>',
            color, obj.get_interpretation_display()
        )
    interpretation_display.short_description = 'Interprétation'
    interpretation_display.admin_order_field = 'interpretation'
    
    def score_details_display(self, obj):
        """Afficher les détails des scores avec poids"""
        if not obj.total_score:
            return 'Aucun score disponible'
        
        details = obj.get_score_details
        html = '<div style="margin: 10px 0;">'
        html += '<h4>Détails des scores pondérés :</h4>'
        html += '<table style="width: 100%; border-collapse: collapse; margin-top: 10px;">'
        html += '<tr style="background-color: #f8f9fa;">'
        html += '<th style="padding: 8px; border: 1px solid #dee2e6; text-align: left;">KPI</th>'
        html += '<th style="padding: 8px; border: 1px solid #dee2e6; text-align: center;">Note</th>'
        html += '<th style="padding: 8px; border: 1px solid #dee2e6; text-align: center;">Poids</th>'
        html += '<th style="padding: 8px; border: 1px solid #dee2e6; text-align: center;">Score Pondéré</th>'
        html += '</tr>'
        
        for kpi_name, data in details.items():
            # Formater le nom du KPI pour l'affichage
            display_name = kpi_name.replace('_', ' ').title()
            if kpi_name == 'delivery_satisfaction_rate':
                display_name = 'Taux de Satisfaction des Livrables'
            elif kpi_name == 'deadline_respect_rate':
                display_name = 'Respect des Délais'
            elif kpi_name == 'learning_capacity':
                display_name = 'Capacité d\'Apprentissage'
            elif kpi_name == 'initiative_taking':
                display_name = 'Prise d\'Initiatives'
            elif kpi_name == 'professional_behavior':
                display_name = 'Comportement Professionnel'
            elif kpi_name == 'adaptability':
                display_name = 'Adaptabilité'
            
            html += '<tr>'
            html += f'<td style="padding: 8px; border: 1px solid #dee2e6;">{display_name}</td>'
            html += f'<td style="padding: 8px; border: 1px solid #dee2e6; text-align: center;">{data["score"]}/5</td>'
            html += f'<td style="padding: 8px; border: 1px solid #dee2e6; text-align: center;">{data["weight"] * 100}%</td>'
            html += f'<td style="padding: 8px; border: 1px solid #dee2e6; text-align: center;">{data["weighted_score"]:.2f}</td>'
            html += '</tr>'
        
        html += '</table>'
        html += f'<p style="margin-top: 15px;"><strong>Score total : {obj.total_score}/5</strong></p>'
        html += '</div>'
        
        return mark_safe(html)
    score_details_display.short_description = 'Détails des Scores'
    
    def weights_summary_display(self, obj):
        """Afficher le résumé des poids"""
        weights = obj.get_weights_summary
        html = '<div style="margin: 10px 0;">'
        html += '<h4>Répartition des poids :</h4>'
        html += '<ul style="list-style-type: none; padding: 0;">'
        
        for kpi_name, weight_info in weights.items():
            # Formater le nom du KPI pour l'affichage
            display_name = kpi_name.replace('_', ' ').title()
            if kpi_name == 'delivery_satisfaction_rate':
                display_name = 'Taux de Satisfaction des Livrables'
            elif kpi_name == 'deadline_respect_rate':
                display_name = 'Respect des Délais'
            elif kpi_name == 'learning_capacity':
                display_name = 'Capacité d\'Apprentissage'
            elif kpi_name == 'initiative_taking':
                display_name = 'Prise d\'Initiatives'
            elif kpi_name == 'professional_behavior':
                display_name = 'Comportement Professionnel'
            elif kpi_name == 'adaptability':
                display_name = 'Adaptabilité'
            
            html += f'<li style="margin: 5px 0; padding: 5px; background-color: #f8f9fa; border-radius: 3px;">'
            html += f'<strong>{display_name}</strong> : {weight_info["percentage"]}'
            html += '</li>'
        
        html += '</ul>'
        html += '<p><strong>Total : 100%</strong></p>'
        html += '</div>'
        
        return mark_safe(html)
    weights_summary_display.short_description = 'Répartition des Poids'
    
    # Actions personnalisées
    actions = ['export_to_excel', 'mark_as_reviewed']
    
    def export_to_excel(self, request, queryset):
        """Exporter les évaluations sélectionnées au format Excel"""
        from django.http import HttpResponse
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from io import BytesIO
        
        # Créer un nouveau classeur Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Évaluations KPI Exportées"
        
        # En-têtes
        headers = [
            'Stagiaire', 'Évaluateur', 'Stage', 'Date d\'évaluation',
            'Taux de Satisfaction des Livrables', 'Respect des Délais',
            'Capacité d\'Apprentissage', 'Prise d\'Initiatives',
            'Comportement Professionnel', 'Adaptabilité',
            'Score Total', 'Interprétation', 'Commentaires'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        
        # Données
        for row, evaluation in enumerate(queryset, 2):
            ws.cell(row=row, column=1, value=evaluation.intern.get_full_name() if evaluation.intern else '')
            ws.cell(row=row, column=2, value=evaluation.evaluator.get_full_name() if evaluation.evaluator else '')
            ws.cell(row=row, column=3, value=evaluation.stage.title if evaluation.stage else '')
            ws.cell(row=row, column=4, value=evaluation.evaluation_date.strftime("%d/%m/%Y") if evaluation.evaluation_date else '')
            ws.cell(row=row, column=5, value=evaluation.delivery_satisfaction_rate)
            ws.cell(row=row, column=6, value=evaluation.deadline_respect_rate)
            ws.cell(row=row, column=7, value=evaluation.learning_capacity)
            ws.cell(row=row, column=8, value=evaluation.initiative_taking)
            ws.cell(row=row, column=9, value=evaluation.professional_behavior)
            ws.cell(row=row, column=10, value=evaluation.adaptability)
            ws.cell(row=row, column=11, value=evaluation.total_score)
            ws.cell(row=row, column=12, value=evaluation.get_interpretation_display())
            ws.cell(row=row, column=13, value=evaluation.comments or '')
        
        # Ajuster la largeur des colonnes
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20
        
        # Créer la réponse HTTP
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="evaluations_kpi_export_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        
        # Sauvegarder dans le buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        response.write(buffer.getvalue())
        
        return response
    export_to_excel.short_description = "Exporter les évaluations sélectionnées en Excel"
    
    def mark_as_reviewed(self, request, queryset):
        """Marquer les évaluations comme révisées"""
        updated = queryset.update(updated_at=timezone.now())
        self.message_user(
            request, 
            f'{updated} évaluation(s) marquée(s) comme révisée(s).'
        )
    mark_as_reviewed.short_description = "Marquer comme révisées"
    
    # Configuration de la pagination et de l'ordre
    list_per_page = 25
    ordering = ['-evaluation_date', '-created_at']
    
    # Configuration des permissions
    def has_add_permission(self, request):
        """Seuls les superusers peuvent ajouter des évaluations via l'admin"""
        return request.user.is_superuser
    
    def has_change_permission(self, request, obj=None):
        """Seuls les superusers peuvent modifier des évaluations via l'admin"""
        return request.user.is_superuser
    
    def has_delete_permission(self, request, obj=None):
        """Seuls les superusers peuvent supprimer des évaluations via l'admin"""
        return request.user.is_superuser
    
    def has_view_permission(self, request, obj=None):
        """Les RH et admins peuvent voir les évaluations"""
        return request.user.role in ['rh', 'admin'] or request.user.is_superuser
