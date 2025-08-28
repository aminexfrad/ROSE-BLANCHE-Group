"""
© 2025 Mohamed Amine FRAD. All rights reserved.
Unauthorized use, reproduction, or modification of this code is strictly prohibited.
Intellectual Property – Protected by international copyright law.
"""

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework import status
from django.shortcuts import get_object_or_404

from django.utils import timezone
from datetime import timedelta, datetime
from django.http import HttpResponse
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from django.db import models

from demande_service.models import Demande as DemandeModel
from shared.models import Stage, Testimonial, Evaluation, Notification, Entreprise
from auth_service.models import User


def _get_excel_styles():
    """Helper function to get common Excel styles"""
    return {
        'header_font': Font(bold=True, color="FFFFFF"),
        'header_fill': PatternFill(start_color="366092", end_color="366092", fill_type="solid"),
        'header_alignment': Alignment(horizontal="center", vertical="center")
    }


def _apply_header_styles(ws, headers):
    """Helper function to apply header styles to worksheet"""
    styles = _get_excel_styles()
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = styles['header_font']
        cell.fill = styles['header_fill']
        cell.alignment = styles['header_alignment']


def _get_kpi_definitions():
    """Helper function to get KPI definitions"""
    return [
        {
            "name": "Taux de Satisfaction des Livrables",
            "definition": "% des livrables validés sans corrections majeures",
            "measurement": "Suivi sur la qualité des travaux",
            "weight": "25%"
        },
        {
            "name": "Respect des Délais",
            "definition": "% des tâches terminées dans les délais prévus",
            "measurement": "Suivi des livrables et des délais",
            "weight": "20%"
        },
        {
            "name": "Capacité d'Apprentissage",
            "definition": "Temps nécessaire pour maîtriser de nouvelles compétences ou outils",
            "measurement": "Temps moyen pour acquérir une nouvelle compétence ou utiliser un nouvel outil",
            "weight": "15%"
        },
        {
            "name": "Prise d'Initiatives",
            "definition": "Nombre d'initiatives ou propositions d'amélioration soumises par le stagiaire",
            "measurement": "Compte des suggestions ou actions prises par le stagiaire pour améliorer le travail ou résoudre un problème",
            "weight": "10%"
        },
        {
            "name": "Comportement en Entreprise et Conduite",
            "definition": "Respect des normes professionnelles, éthique, communication",
            "measurement": "Observation directe sur la capacité du stagiaire à respecter les règles, maintenir une éthique professionnelle et communiquer efficacement avec les autres",
            "weight": "15%"
        },
        {
            "name": "Adaptabilité au Changement",
            "definition": "Capacité à s'adapter aux changements de tâches ou de priorités",
            "measurement": "Observation et feedbacks des managers sur la flexibilité du stagiaire",
            "weight": "15%"
        }
    ]


def _set_column_widths(ws, num_columns=6, width=25):
    """Helper function to set column widths"""
    for col in range(1, num_columns + 1):
        ws.column_dimensions[get_column_letter(col)].width = width


def _create_excel_response(wb, filename):
    """Helper function to create Excel HTTP response"""
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response.write(buffer.getvalue())
    
    return response


def _get_excel_headers():
    """Helper function to get Excel headers"""
    return [
        "KPI", "Définition", "Méthode de Mesure", "Poids", 
        "Note (sur 5)", "Score Calculé"
    ]


def _create_excel_workbook(title):
    """Helper function to create Excel workbook with title"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title
    return wb, ws


def get_company_filtered_queryset(request, base_queryset, company_field='company_entreprise'):
    """
    Utility function to filter queryset based on RH user's company access.
    RH users can only see data from their assigned company.
    Admin users can see all data.
    """
    if request.user.role == 'rh':
        if not request.user.entreprise:
            # RH users without company assignment see no data
            return base_queryset.none()
        
        # Filter by company
        if company_field == 'company_entreprise':
            return base_queryset.filter(company_entreprise=request.user.entreprise)
        elif company_field == 'entreprise':
            return base_queryset.filter(entreprise=request.user.entreprise)
        elif company_field == 'stagiaire__entreprise':
            return base_queryset.filter(stagiaire__entreprise=request.user.entreprise)
        elif company_field == 'evaluated__entreprise':
            return base_queryset.filter(evaluated__entreprise=request.user.entreprise)
        elif company_field == 'author__entreprise':
            return base_queryset.filter(author__entreprise=request.user.entreprise)
        else:
            # Default company filtering
            return base_queryset.filter(**{company_field: request.user.entreprise})
    
    elif request.user.role == 'admin':
        # Admin can see all data
        return base_queryset
    
    else:
        # Other roles see no data
        return base_queryset.none()


def validate_rh_company_access(request, target_entreprise):
    """
    Validate that RH user has access to the target company.
    Returns (has_access, error_message)
    """
    if request.user.role == 'admin':
        return True, None
    
    if request.user.role != 'rh':
        return False, "Permission refusée - rôle non autorisé"
    
    if not request.user.entreprise:
        return False, "Permission refusée - utilisateur RH non associé à une entreprise"
    
    if request.user.entreprise != target_entreprise:
        return False, "Permission refusée - accès à cette entreprise non autorisé"
    
    return True, None


class RHStagiairesView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        try:
            # Use company-based filtering
            stagiaires = get_company_filtered_queryset(
                request, 
                User.objects.filter(role='stagiaire'),
                'entreprise'
            ).order_by('-date_joined')
            
            stagiaires_data = []
            for stagiaire in stagiaires:
                # Get their active stage
                active_stage = Stage.objects.filter(stagiaire=stagiaire, status='active').first()
                
                stagiaire_data = {
                    "id": stagiaire.id,
                    "prenom": stagiaire.prenom,
                    "nom": stagiaire.nom,
                    "first_name": stagiaire.prenom,
                    "last_name": stagiaire.nom,
                    "email": stagiaire.email,
                    "telephone": stagiaire.telephone,
                    "institut": stagiaire.institut,
                    "specialite": stagiaire.specialite,
                    "avatar": stagiaire.avatar.url if stagiaire.avatar else None,
                    "date_joined": stagiaire.date_joined.isoformat(),
                    "entreprise": stagiaire.entreprise.nom if stagiaire.entreprise else None,
                    "active_stage": {
                        "id": active_stage.id,
                        "title": active_stage.title,
                        "company": active_stage.company_entreprise.nom if active_stage.company_entreprise else active_stage.company_name,
                        "progress": active_stage.progress,
                        "start_date": active_stage.start_date.isoformat(),
                        "end_date": active_stage.end_date.isoformat(),
                        "tuteur": {
                            "id": active_stage.tuteur.id,
                            "prenom": active_stage.tuteur.prenom,
                            "nom": active_stage.tuteur.nom,
                            "first_name": active_stage.tuteur.prenom,
                            "last_name": active_stage.tuteur.nom,
                            "email": active_stage.tuteur.email
                        } if active_stage.tuteur else None
                    } if active_stage else None
                }
                stagiaires_data.append(stagiaire_data)
            
            return Response({
                "results": stagiaires_data
            })
            
        except Exception as e:
            return Response(
                {'error': f'Error fetching stagiaires data: {str(e)}'}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

class RHStagiaireDetailView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, pk):
        try:
            # Get the stagiaire
            stagiaire = get_object_or_404(User, id=pk, role='stagiaire')
            
            # Validate company access for RH users
            if request.user.role == 'rh':
                has_access, error_msg = validate_rh_company_access(
                    request, 
                    stagiaire.entreprise
                )
                if not has_access:
                    return Response(
                        {'error': error_msg}, 
                        status=status.HTTP_403_FORBIDDEN
                    )
            
            # Get their stages
            stages = Stage.objects.filter(stagiaire=stagiaire).order_by('-created_at')
            
            stages_data = []
            for stage in stages:
                # Get steps for this stage
                steps = stage.steps.all().order_by('order')
                total_steps = steps.count()
                completed_steps = steps.filter(status__in=['completed', 'validated']).count()
                progress = int((completed_steps / total_steps * 100) if total_steps > 0 else 0)
                
                stage_data = {
                    "id": stage.id,
                    "title": stage.title,
                    "company": stage.company_entreprise.nom if stage.company_entreprise else stage.company_name,
                    "status": stage.status,
                    "start_date": stage.start_date.isoformat(),
                    "end_date": stage.end_date.isoformat(),
                    "progress": progress,
                    "location": stage.location,
                    "description": stage.description,
                    "tuteur": {
                        "id": stage.tuteur.id,
                        "prenom": stage.tuteur.prenom,
                        "nom": stage.tuteur.nom,
                        "first_name": stage.tuteur.prenom,
                        "last_name": stage.tuteur.nom,
                        "email": stage.tuteur.email
                    } if stage.tuteur else None,
                    "steps_count": total_steps,
                    "completed_steps_count": completed_steps,
                    "documents_count": stage.documents.count(),
                    "evaluations_count": stage.evaluations.count()
                }
                stages_data.append(stage_data)
            
            response_data = {
                "stagiaire": {
                    "id": stagiaire.id,
                    "prenom": stagiaire.prenom,
                    "nom": stagiaire.nom,
                    "first_name": stagiaire.prenom,
                    "last_name": stagiaire.nom,
                    "email": stagiaire.email,
                    "telephone": stagiaire.telephone,
                    "institut": stagiaire.institut,
                    "specialite": stagiaire.specialite,
                    "departement": stagiaire.departement,
                    "avatar": stagiaire.avatar.url if stagiaire.avatar else None,
                    "date_joined": stagiaire.date_joined.isoformat(),
                    "last_login": stagiaire.last_login.isoformat() if stagiaire.last_login else None
                },
                "stages": stages_data
            }
            
            return Response(response_data)
            
        except User.DoesNotExist:
            return Response(
                {'error': 'Stagiaire not found'}, 
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            return Response(
                {'error': f'Error fetching stagiaire data: {str(e)}'}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

class RHTestimonialsView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        try:
            # Use company-based filtering for testimonials
            testimonials = get_company_filtered_queryset(
                request, 
                Testimonial.objects.all(),
                'stage__company_entreprise'
            ).order_by('-created_at')
            
            testimonials_data = []
            for testimonial in testimonials:
                testimonial_data = {
                    "id": testimonial.id,
                    "title": testimonial.title,
                    "content": testimonial.content,
                    "testimonial_type": testimonial.testimonial_type,
                    "status": testimonial.status,
                    "video_url": testimonial.video_url,
                    "video_file": request.build_absolute_uri(testimonial.video_file.url) if testimonial.video_file and testimonial.video_file.name else None,
                    "created_at": testimonial.created_at.isoformat(),
                    "moderated_at": testimonial.moderated_at.isoformat() if testimonial.moderated_at else None,
                    "moderation_comment": testimonial.moderation_comment,
                    "author": {
                        "id": testimonial.author.id,
                        "prenom": testimonial.author.prenom,
                        "nom": testimonial.author.nom,
                        "first_name": testimonial.author.prenom,
                        "last_name": testimonial.author.nom,
                        "email": testimonial.author.email
                    },
                    "stage": {
                        "id": testimonial.stage.id,
                        "title": testimonial.stage.title,
                        "company": testimonial.stage.company_entreprise.nom if testimonial.stage.company_entreprise else testimonial.stage.company_name
                    },
                    "moderated_by": {
                        "id": testimonial.moderated_by.id,
                        "prenom": testimonial.moderated_by.prenom,
                        "nom": testimonial.moderated_by.nom,
                        "first_name": testimonial.moderated_by.prenom,
                        "last_name": testimonial.moderated_by.nom
                    } if testimonial.moderated_by else None
                }
                testimonials_data.append(testimonial_data)
            
            return Response({
                "results": testimonials_data
            })
            
        except Exception as e:
            return Response(
                {'error': f'Error fetching testimonials data: {str(e)}'}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

class RHTestimonialModerationView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, pk):
        return self.moderate_testimonial(request, pk)

    def put(self, request, pk):
        return self.moderate_testimonial(request, pk)

    def moderate_testimonial(self, request, pk):
        try:
            testimonial = get_object_or_404(Testimonial, id=pk)
            
            # Validate company access for RH users
            if request.user.role == 'rh':
                has_access, error_msg = validate_rh_company_access(
                    request, 
                    testimonial.stage.company_entreprise
                )
                if not has_access:
                    return Response(
                        {'error': error_msg}, 
                        status=status.HTTP_403_FORBIDDEN
                    )
            
            action = request.data.get('action')
            comment = request.data.get('comment', '')
            
            if action == 'approve':
                testimonial.status = 'approved'
                testimonial.moderated_by = request.user
                testimonial.moderated_at = timezone.now()
                testimonial.moderation_comment = comment
                testimonial.save()
                
                # Create notification for the author
                Notification.objects.create(
                    recipient=testimonial.author,
                    title='Témoignage approuvé',
                    message=f'Votre témoignage "{testimonial.title}" a été approuvé et publié sur la plateforme.',
                    notification_type='success'
                )
                
                return Response({
                    'message': 'Témoignage approuvé avec succès',
                    'testimonial': {
                        'id': testimonial.id,
                        'status': testimonial.status,
                        'moderated_at': testimonial.moderated_at.isoformat(),
                        'moderation_comment': testimonial.moderation_comment
                    }
                })
                
            elif action == 'reject':
                testimonial.status = 'rejected'
                testimonial.moderated_by = request.user
                testimonial.moderated_at = timezone.now()
                testimonial.moderation_comment = comment
                testimonial.save()
                
                # Create notification for the author
                rejection_message = f'Votre témoignage "{testimonial.title}" nécessite des modifications.'
                if comment:
                    rejection_message += f' Commentaire: {comment}'
                
                Notification.objects.create(
                    recipient=testimonial.author,
                    title='Témoignage nécessite des modifications',
                    message=rejection_message,
                    notification_type='warning'
                )
                
                return Response({
                    'message': 'Témoignage rejeté',
                    'testimonial': {
                        'id': testimonial.id,
                        'status': testimonial.status,
                        'moderated_at': testimonial.moderated_at.isoformat(),
                        'moderation_comment': testimonial.moderation_comment
                    }
                })
                
            else:
                return Response(
                    {'error': 'Action invalide. Utilisez "approve" ou "reject"'}, 
                    status=status.HTTP_400_BAD_REQUEST
                )
                
        except Testimonial.DoesNotExist:
            return Response(
                {'error': 'Témoignage non trouvé'}, 
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            return Response(
                {'error': f'Erreur lors de la modération du témoignage: {str(e)}'}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )



class RHStagesView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        try:
            # Use company-based filtering with optimized queries
            stages = get_company_filtered_queryset(
                request, 
                Stage.objects.select_related('stagiaire', 'tuteur', 'company_entreprise')
                .prefetch_related('steps', 'documents', 'evaluations'),
                'company_entreprise'
            ).order_by('-created_at')
            
            stages_data = []
            for stage in stages:
                # Calculate progress
                total_steps = stage.steps.count()
                completed_steps = stage.steps.filter(status__in=['completed', 'validated']).count()
                progress = int((completed_steps / total_steps * 100) if total_steps > 0 else 0)
                
                stage_data = {
                    "id": stage.id,
                    "title": stage.title,
                    "company": stage.company_entreprise.nom if stage.company_entreprise else stage.company_name,
                    "status": stage.status,
                    "start_date": stage.start_date.isoformat(),
                    "end_date": stage.end_date.isoformat(),
                    "progress": progress,
                    "location": stage.location,
                    "description": stage.description,
                    "stagiaire": {
                        "id": stage.stagiaire.id,
                        "prenom": stage.stagiaire.prenom,
                        "nom": stage.stagiaire.nom,
                        "first_name": stage.stagiaire.prenom,
                        "last_name": stage.stagiaire.nom,
                        "email": stage.stagiaire.email,
                        "institut": stage.stagiaire.institut,
                        "specialite": stage.stagiaire.specialite
                    },
                    "tuteur": {
                        "id": stage.tuteur.id,
                        "prenom": stage.tuteur.prenom,
                        "nom": stage.tuteur.nom,
                        "first_name": stage.tuteur.prenom,
                        "last_name": stage.tuteur.nom,
                        "email": stage.tuteur.email
                    } if stage.tuteur else None,
                    "steps_count": total_steps,
                    "completed_steps_count": completed_steps,
                    "documents_count": stage.documents.count(),
                    "evaluations_count": stage.evaluations.count(),
                    "created_at": stage.created_at.isoformat()
                }
                stages_data.append(stage_data)
            
            return Response({
                "results": stages_data
            })
            
        except Exception as e:
            return Response(
                {'error': f'Error fetching stages data: {str(e)}'}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

class RHStageDetailView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, pk):
        try:
            stage = get_object_or_404(
                Stage.objects.select_related('stagiaire', 'tuteur', 'company_entreprise')
                .prefetch_related('steps', 'documents', 'evaluations'), 
                id=pk
            )
            
            # Validate company access for RH users
            if request.user.role == 'rh':
                has_access, error_msg = validate_rh_company_access(
                    request, 
                    stage.company_entreprise
                )
                if not has_access:
                    return Response(
                        {'error': error_msg}, 
                        status=status.HTTP_403_FORBIDDEN
                    )
            
            # Get steps (already prefetched)
            steps = stage.steps.all().order_by('order')
            steps_data = []
            for step in steps:
                step_data = {
                    "id": step.id,
                    "title": step.title,
                    "description": step.description,
                    "order": step.order,
                    "status": step.status,
                    "due_date": step.due_date.isoformat() if step.due_date else None,
                    "completed_date": step.completed_date.isoformat() if step.completed_date else None,
                    "validated_date": step.validated_date.isoformat() if step.validated_date else None,
                    "tuteur_feedback": step.tuteur_feedback,
                    "stagiaire_comment": step.stagiaire_comment,
                    "documents_count": step.documents.count()
                }
                steps_data.append(step_data)
            
            # Get documents
            documents = stage.documents.all().order_by('-created_at')
            documents_data = []
            for doc in documents:
                doc_data = {
                    "id": doc.id,
                    "title": doc.title,
                    "description": doc.description,
                    "document_type": doc.document_type,
                    "file_url": doc.file.url if doc.file else None,
                    "file_size": doc.file_size,
                    "is_approved": doc.is_approved,
                    "feedback": doc.feedback,
                    "uploaded_at": doc.created_at.isoformat(),
                    "approved_at": doc.approved_at.isoformat() if doc.approved_at else None,
                    "uploaded_by": {
                        "id": doc.uploaded_by.id,
                        "name": doc.uploaded_by.get_full_name(),
                        "role": doc.uploaded_by.role
                    }
                }
                documents_data.append(doc_data)
            
            # Get evaluations
            evaluations = stage.evaluations.all().order_by('-created_at')
            evaluations_data = []
            for eval in evaluations:
                eval_data = {
                    "id": eval.id,
                    "evaluation_type": eval.evaluation_type,
                    "scores": eval.scores,
                    "comments": eval.comments,
                    "overall_score": eval.overall_score,
                    "is_completed": eval.is_completed,
                    "completed_at": eval.completed_at.isoformat() if eval.completed_at else None,
                    "evaluator": {
                        "id": eval.evaluator.id,
                        "name": eval.evaluator.get_full_name(),
                        "role": eval.evaluator.role
                    },
                    "evaluated": {
                        "id": eval.evaluated.id,
                        "name": eval.evaluated.get_full_name(),
                        "role": eval.evaluated.role
                    },
                    "created_at": eval.created_at.isoformat()
                }
                evaluations_data.append(eval_data)
            
            # Calculate overall progress
            total_steps = len(steps)
            completed_steps = len([s for s in steps if s.status in ['completed', 'validated']])
            progress = int((completed_steps / total_steps * 100) if total_steps > 0 else 0)
            
            response_data = {
                "stage": {
                    "id": stage.id,
                    "title": stage.title,
                    "company": stage.company_entreprise.nom if stage.company_entreprise else stage.company_name,
                    "status": stage.status,
                    "start_date": stage.start_date.isoformat(),
                    "end_date": stage.end_date.isoformat(),
                    "progress": progress,
                    "location": stage.location,
                    "description": stage.description,
                    "created_at": stage.created_at.isoformat()
                },
                "stagiaire": {
                    "id": stage.stagiaire.id,
                    "first_name": stage.stagiaire.prenom,
                    "last_name": stage.stagiaire.nom,
                    "email": stage.stagiaire.email,
                    "telephone": stage.stagiaire.telephone,
                    "institut": stage.stagiaire.institut,
                    "specialite": stage.stagiaire.specialite,
                    "avatar": stage.stagiaire.avatar.url if stage.stagiaire.avatar else None
                },
                "tuteur": {
                    "id": stage.tuteur.id,
                    "first_name": stage.tuteur.prenom,
                    "last_name": stage.tuteur.nom,
                    "email": stage.tuteur.email,
                    "telephone": stage.tuteur.telephone
                } if stage.tuteur else None,
                "steps": steps_data,
                "documents": documents_data,
                "evaluations": evaluations_data
            }
            
            return Response(response_data)
            
        except Stage.DoesNotExist:
            return Response(
                {'error': 'Stage not found'}, 
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            return Response(
                {'error': f'Error fetching stage data: {str(e)}'}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

class RHEvaluationsView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        try:
            # Use company-based filtering for evaluations
            evaluations = get_company_filtered_queryset(
                request, 
                Evaluation.objects.all(),
                'stage__company_entreprise'
            ).order_by('-created_at')
            
            evaluations_data = []
            for eval in evaluations:
                eval_data = {
                    "id": eval.id,
                    "evaluation_type": eval.evaluation_type,
                    "scores": eval.scores,
                    "comments": eval.comments,
                    "overall_score": eval.overall_score,
                    "is_completed": eval.is_completed,
                    "completed_at": eval.completed_at.isoformat() if eval.completed_at else None,
                    "evaluator": {
                        "id": eval.evaluator.id,
                        "name": eval.evaluator.get_full_name(),
                        "role": eval.evaluator.role
                    },
                    "evaluated": {
                        "id": eval.evaluated.id,
                        "name": eval.evaluated.get_full_name(),
                        "role": eval.evaluated.role
                    },
                    "stage": {
                        "id": eval.stage.id,
                        "title": eval.stage.title,
                        "company": eval.stage.company_entreprise.nom if eval.stage.company_entreprise else eval.stage.company_name
                    },
                    "created_at": eval.created_at.isoformat()
                }
                evaluations_data.append(eval_data)
            
            return Response({
                "results": evaluations_data
            })
            
        except Exception as e:
            return Response(
                {'error': f'Error fetching evaluations data: {str(e)}'}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

class RHNotificationsView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        try:
            # Use company-based filtering for notifications
            notifications = get_company_filtered_queryset(
                request, 
                Notification.objects.all(),
                'recipient__entreprise'
            ).order_by('-created_at')
            
            notifications_data = []
            for notification in notifications:
                notification_data = {
                    "id": notification.id,
                    "title": notification.title,
                    "message": notification.message,
                    "notification_type": notification.notification_type,
                    "is_read": notification.is_read,
                    "read_at": notification.read_at.isoformat() if notification.read_at else None,
                    "recipient": {
                        "id": notification.recipient.id,
                        "name": notification.recipient.get_full_name(),
                        "role": notification.recipient.role
                    },
                    "created_at": notification.created_at.isoformat()
                }
                notifications_data.append(notification_data)
            
            return Response({
                "results": notifications_data
            })
            
        except Exception as e:
            return Response(
                {'error': f'Error fetching notifications data: {str(e)}'}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )







class RHTuteursDisponiblesView(APIView):
    """
    Vue pour récupérer la liste des tuteurs disponibles
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        try:
            # Vérifier que l'utilisateur est RH ou admin
            if request.user.role not in ['rh', 'admin']:
                return Response(
                    {'error': 'Permission refusée'}, 
                    status=status.HTTP_403_FORBIDDEN
                )
            
            # Use company-based filtering
            tuteurs = get_company_filtered_queryset(
                request, 
                User.objects.filter(role='tuteur'),
                'entreprise'
            ).order_by('prenom', 'nom')
            
            tuteurs_data = []
            for tuteur in tuteurs:
                # Compter le nombre de stagiaires actuellement assignés
                stagiaires_assignes = Stage.objects.filter(
                    tuteur=tuteur, 
                    status='active'
                ).count()
                
                tuteur_data = {
                    "id": tuteur.id,
                    "first_name": tuteur.prenom,
                    "last_name": tuteur.nom,
                    "email": tuteur.email,
                    "telephone": tuteur.telephone,
                    "departement": tuteur.departement,
                    "stagiaires_assignes": stagiaires_assignes,
                    "disponible": stagiaires_assignes < 5,  # Limite de 5 stagiaires par tuteur
                    "entreprise": tuteur.entreprise.nom if tuteur.entreprise else None
                }
                tuteurs_data.append(tuteur_data)
            
            return Response({
                "results": tuteurs_data,
                "count": len(tuteurs_data)
            })
            
        except Exception as e:
            return Response(
                {'error': f'Erreur lors de la récupération des tuteurs: {str(e)}'}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

class RHAssignerTuteurView(APIView):
    """
    Vue pour assigner un tuteur à un stagiaire
    """
    permission_classes = [IsAuthenticated]

    def post(self, request, pk):
        try:
            # Vérifier que l'utilisateur est RH ou admin
            if request.user.role not in ['rh', 'admin']:
                return Response(
                    {'error': 'Permission refusée'}, 
                    status=status.HTTP_403_FORBIDDEN
                )
            
            # Récupérer le stagiaire
            stagiaire = get_object_or_404(User, id=pk, role='stagiaire')
            
            # Validate company access for RH users
            if request.user.role == 'rh':
                has_access, error_msg = validate_rh_company_access(
                    request, 
                    stagiaire.entreprise
                )
                if not has_access:
                    return Response(
                        {'error': error_msg}, 
                        status=status.HTTP_403_FORBIDDEN
                    )
            
            # Récupérer l'ID du tuteur depuis la requête
            tuteur_id = request.data.get('tuteur_id')
            if not tuteur_id:
                return Response(
                    {'error': 'ID du tuteur requis'}, 
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Récupérer le tuteur
            tuteur = get_object_or_404(User, id=tuteur_id, role='tuteur')
            
            # Vérifier que le tuteur n'a pas trop de stagiaires assignés
            stagiaires_assignes = Stage.objects.filter(
                tuteur=tuteur, 
                status='active'
            ).count()
            
            if stagiaires_assignes >= 5:
                return Response(
                    {'error': 'Ce tuteur a déjà le maximum de stagiaires assignés (5)'}, 
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Vérifier que le tuteur appartient à la même entreprise que le stagiaire
            if stagiaire.entreprise != tuteur.entreprise:
                return Response(
                    {'error': 'Le tuteur sélectionné n\'appartient pas à la même entreprise que le stagiaire.'}, 
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Récupérer le stage actif du stagiaire
            stage_actif = Stage.objects.filter(
                stagiaire=stagiaire, 
                status='active'
            ).first()
            
            if not stage_actif:
                return Response(
                    {'error': 'Aucun stage actif trouvé pour ce stagiaire'}, 
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Assigner le tuteur au stage
            stage_actif.tuteur = tuteur
            stage_actif.save()
            
            # Créer une notification pour le tuteur
            Notification.objects.create(
                recipient=tuteur,
                title="Nouveau stagiaire assigné",
                message=f"Vous avez été assigné comme tuteur pour {stagiaire.prenom} {stagiaire.nom}",
                notification_type="info",
                related_stage=stage_actif
            )
            
            # Créer une notification pour le stagiaire
            Notification.objects.create(
                recipient=stagiaire,
                title="Tuteur assigné",
                message=f"Un tuteur a été assigné à votre stage : {tuteur.prenom} {tuteur.nom}",
                notification_type="success",
                related_stage=stage_actif
            )
            
            return Response({
                'message': 'Tuteur assigné avec succès',
                'stagiaire': {
                    'id': stagiaire.id,
                    'first_name': stagiaire.prenom,
                    'last_name': stagiaire.nom,
                    'email': stagiaire.email
                },
                'tuteur': {
                    'id': tuteur.id,
                    'first_name': tuteur.prenom,
                    'last_name': tuteur.nom,
                    'email': tuteur.email
                },
                'stage': {
                    'id': stage_actif.id,
                    'title': stage_actif.title,
                    'company': (stage_actif.company_entreprise.nom if stage_actif.company_entreprise else stage_actif.company_name)
                }
            }, status=status.HTTP_200_OK)
            
        except User.DoesNotExist:
            return Response(
                {'error': 'Utilisateur non trouvé'}, 
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            return Response(
                {'error': f'Erreur lors de l\'assignation: {str(e)}'}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

class RHCreateStageForStagiaireView(APIView):
    """
    Vue pour créer un stage pour un stagiaire existant
    """
    permission_classes = [IsAuthenticated]

    def post(self, request, pk):
        try:
            # Vérifier que l'utilisateur est RH ou admin
            if request.user.role not in ['rh', 'admin']:
                return Response(
                    {'error': 'Permission refusée'}, 
                    status=status.HTTP_403_FORBIDDEN
                )
            
            # Récupérer le stagiaire
            stagiaire = get_object_or_404(User, id=pk, role='stagiaire')
            
            # Validate company access for RH users
            if request.user.role == 'rh':
                has_access, error_msg = validate_rh_company_access(
                    request, 
                    stagiaire.entreprise
                )
                if not has_access:
                    return Response(
                        {'error': error_msg}, 
                        status=status.HTTP_403_FORBIDDEN
                    )
            
            # Vérifier s'il a déjà un stage actif
            existing_stage = Stage.objects.filter(
                stagiaire=stagiaire, 
                status='active'
            ).first()
            
            if existing_stage:
                return Response({
                    'message': 'Ce stagiaire a déjà un stage actif',
                    'stage': {
                        'id': existing_stage.id,
                        'title': existing_stage.title,
                        'status': existing_stage.status
                    }
                }, status=status.HTTP_200_OK)
            
            # Récupérer les données du stage
            data = request.data
            required_fields = ['title', 'company', 'location', 'start_date', 'end_date']
            for field in required_fields:
                if not data.get(field):
                    return Response(
                        {'error': f'Le champ {field} est obligatoire'}, 
                        status=status.HTTP_400_BAD_REQUEST
                    )
            
            # Déterminer l'entreprise pour le stage
            # Si l'utilisateur RH a une entreprise, l'utiliser
            # Sinon, essayer de trouver l'entreprise par le nom
            entreprise = None
            if request.user.role == 'rh' and request.user.entreprise:
                entreprise = request.user.entreprise
            else:
                # Essayer de trouver l'entreprise par le nom
                try:
                    entreprise = Entreprise.objects.get(nom=data['company'])
                except Entreprise.DoesNotExist:
                    # Créer une entreprise temporaire si elle n'existe pas
                    entreprise = Entreprise.objects.create(
                        nom=data['company'],
                        description=f"Entreprise créée automatiquement pour {stagiaire.prenom} {stagiaire.nom}",
                        secteur_activite="Non spécifié"
                    )
            
            # Créer une demande de stage approuvée
            demande = DemandeModel.objects.create(
                nom=stagiaire.nom,
                prenom=stagiaire.prenom,
                email=stagiaire.email,
                telephone=stagiaire.telephone or '',
                cin=f"CIN{stagiaire.id:06d}",
                institut=stagiaire.institut or '',
                specialite=stagiaire.specialite or '',
                niveau=data.get('niveau', 'Bac+3'),
                type_stage=data.get('type_stage', 'Stage PFE'),
                date_debut=data['start_date'],
                date_fin=data['end_date'],
                stage_binome=False,
                status='approved',
                user_created=stagiaire,
                entreprise=entreprise
            )
            
            # Créer le stage
            stage = Stage.objects.create(
                demande=demande,
                stagiaire=stagiaire,
                title=data['title'],
                description=data.get('description', ''),
                company_entreprise=entreprise,
                company_name=data['company'],  # Garder pour la compatibilité
                location=data['location'],
                start_date=data['start_date'],
                end_date=data['end_date'],
                status='active',
                progress=0
            )
            
            return Response({
                'message': 'Stage créé avec succès',
                'stage': {
                    'id': stage.id,
                    'title': stage.title,
                    'company': stage.company_entreprise.nom if stage.company_entreprise else stage.company_name,
                    'location': stage.location,
                    'start_date': stage.start_date,
                    'end_date': stage.end_date,
                    'status': stage.status
                }
            }, status=status.HTTP_201_CREATED)
            
        except User.DoesNotExist:
            return Response(
                {'error': 'Stagiaire non trouvé'}, 
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            return Response(
                {'error': f'Erreur lors de la création du stage: {str(e)}'}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )





# ============================================================================
# VUES POUR LES ÉVALUATIONS KPI DES STAGIAIRES
# ============================================================================

from rest_framework import viewsets, permissions
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.decorators import action
from .models import InternKpiEvaluation
from .serializers import (
    InternKpiEvaluationSerializer,
    InternKpiEvaluationCreateSerializer,
    InternKpiEvaluationUpdateSerializer,
    InternKpiEvaluationListSerializer,
    InternKpiEvaluationDetailSerializer
)
from .permissions import IsRHUser

class InternKpiEvaluationViewSet(viewsets.ModelViewSet):
    """
    ViewSet pour la gestion des évaluations KPI des stagiaires
    """
    queryset = InternKpiEvaluation.objects.all()
    permission_classes = [permissions.IsAuthenticated, IsRHUser]
    parser_classes = [MultiPartParser, FormParser]
    
    def get_serializer_class(self):
        """Retourner le serializer approprié selon l'action"""
        if self.action == 'create':
            return InternKpiEvaluationCreateSerializer
        elif self.action in ['update', 'partial_update']:
            return InternKpiEvaluationUpdateSerializer
        elif self.action == 'list':
            return InternKpiEvaluationListSerializer
        elif self.action == 'retrieve':
            return InternKpiEvaluationDetailSerializer
        return InternKpiEvaluationSerializer
    
    def get_queryset(self):
        """Filtrer les évaluations selon les permissions et paramètres"""
        queryset = super().get_queryset()
        
        # Filtrer par stagiaire si spécifié
        intern_id = self.request.query_params.get('intern_id')
        if intern_id:
            queryset = queryset.filter(intern_id=intern_id)
        
        # Filtrer par stage si spécifié
        stage_id = self.request.query_params.get('stage_id')
        if stage_id:
            queryset = queryset.filter(stage_id=stage_id)
        
        # Filtrer par date d'évaluation
        start_date = self.request.query_params.get('start_date')
        end_date = self.request.query_params.get('end_date')
        if start_date:
            queryset = queryset.filter(evaluation_date__gte=start_date)
        if end_date:
            queryset = queryset.filter(evaluation_date__lte=end_date)
        
        # Filtrer par catégorie de potentiel
        interpretation = self.request.query_params.get('interpretation')
        if interpretation:
            queryset = queryset.filter(interpretation=interpretation)
        
        # Filtrer par score total
        min_score = self.request.query_params.get('min_score')
        max_score = self.request.query_params.get('max_score')
        if min_score:
            queryset = queryset.filter(total_score__gte=min_score)
        if max_score:
            queryset = queryset.filter(total_score__lte=max_score)
        
        return queryset.select_related('intern', 'evaluator', 'stage')
    
    def perform_create(self, serializer):
        """Créer l'évaluation avec l'évaluateur automatiquement"""
        serializer.save(evaluator=self.request.user)
    
    @action(detail=False, methods=['get'])
    def statistics(self, request):
        """Récupérer les statistiques des évaluations KPI"""
        queryset = self.get_queryset()
        
        # Statistiques générales
        total_evaluations = queryset.count()
        if total_evaluations == 0:
            return Response({
                'total_evaluations': 0,
                'average_score': 0,
                'interpretation_distribution': {},
                'recent_evaluations': 0
            })
        
        # Score moyen
        average_score = queryset.aggregate(
            avg_score=models.Avg('total_score')
        )['avg_score']
        
        # Distribution des interprétations
        interpretation_distribution = {}
        for choice in InternKpiEvaluation.PotentialCategory.choices:
            count = queryset.filter(interpretation=choice[0]).count()
            interpretation_distribution[choice[1]] = count
        
        # Évaluations récentes (30 derniers jours)
        thirty_days_ago = timezone.now().date() - timedelta(days=30)
        recent_evaluations = queryset.filter(
            evaluation_date__gte=thirty_days_ago
        ).count()
        
        # Top 5 des meilleurs scores
        top_scores = queryset.order_by('-total_score')[:5].values(
            'intern__nom', 'intern__prenom', 'total_score', 'interpretation'
        )
        
        return Response({
            'total_evaluations': total_evaluations,
            'average_score': round(average_score, 2),
            'interpretation_distribution': interpretation_distribution,
            'recent_evaluations': recent_evaluations,
            'top_scores': top_scores
        })
    
    @action(detail=False, methods=['get'])
    def export_excel(self, request):
        """Exporter les évaluations au format Excel"""
        queryset = self.get_queryset()
        
        # Créer un nouveau classeur Excel
        wb, ws = _create_excel_workbook("Évaluations KPI Stagiaires")
        
        # En-têtes
        headers = _get_excel_headers()
        _apply_header_styles(ws, headers)
        
        # Définitions des KPIs
        kpi_definitions = _get_kpi_definitions()
        
        # Remplir les données des KPIs
        for row, kpi in enumerate(kpi_definitions, 2):
            ws.cell(row=row, column=1, value=kpi["name"])
            ws.cell(row=row, column=2, value=kpi["definition"])
            ws.cell(row=row, column=3, value=kpi["measurement"])
            ws.cell(row=row, column=4, value=kpi["weight"])
            ws.cell(row=row, column=5, value="")  # Note à remplir
            ws.cell(row=row, column=6, value="0")  # Score calculé
        
        # Ligne du total
        total_row = len(kpi_definitions) + 2
        ws.cell(row=total_row, column=1, value="Total des points (sur 100)")
        ws.cell(row=total_row, column=2, value="")
        ws.cell(row=total_row, column=3, value="")
        ws.cell(row=total_row, column=4, value="100%")
        ws.cell(row=total_row, column=5, value="")
        ws.cell(row=total_row, column=6, value="0/5")
        
        # Tableau d'interprétation
        interpretation_row = total_row + 2
        ws.cell(row=interpretation_row, column=1, value="3")
        ws.cell(row=interpretation_row, column=2, value="Catégorie")
        
        interpretation_data = [
            ["Potentiel élevé", "De 4,5 à 5"],
            ["Bon potentiel", "De 3,5 à 4,4"],
            ["Potentiel moyen", "De 2,5 à 3,4"],
            ["Potentiel à renforcer", "En dessous de 2,5"]
        ]
        
        for i, (category, range_desc) in enumerate(interpretation_data):
            ws.cell(row=interpretation_row + i + 1, column=1, value="")
            ws.cell(row=interpretation_row + i + 1, column=2, value=category)
            ws.cell(row=interpretation_row + i + 1, column=3, value=range_desc)
        
        # Ajuster la largeur des colonnes
        _set_column_widths(ws)
        
        # Créer la réponse HTTP
        filename = f"evaluations_kpi_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return _create_excel_response(wb, filename)
    

